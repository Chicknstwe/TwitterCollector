# -*- coding: utf-8 -*-

class TwitterCollector():
    
    def __init__(self):
        
        import json
        import twitter
        
        self.CONFIG = {
                # Use last ids to collect tweets and save a record for later use
                'use_last_ids': True,
                
                # Info and stats configuration
                'show_sleep': False,
                'show_info': False,
                'show_tweet_quantity': True,
                
                # WARNING: this may cause the code to reach the API rate limit
                'show_rate_limit': False,
                
                'get_users_mode': 'file',
                
                # Excel export config
                'auto_export': False,
                'export_mode': 'split'
                }            
            
        try:
            with open('oauth.json', 'r') as file:
                self.oauth = json.load(file)
                
        except:
            self.generateOauth()
            with open('oauth.json', 'r') as file:
                self.oauth = json.load(file)
                
        self.api = twitter.Api(consumer_key=self.oauth['consumer_key'],
                  consumer_secret=self.oauth['consumer_secret'],
                  access_token_key=self.oauth['access_token_key'],
                  access_token_secret=self.oauth['access_token_secret'],
                  tweet_mode='extended',
                  sleep_on_rate_limit=True)        

        
    def generateOauth(self):
        
        import json
        
        tokens = {'consumer_key':None, 'consumer_secret':None, 'access_token_key':None, 'access_token_secret':None}
        
        print('No existe archivo de autoridad o está corrupto. Se creará uno nuevo.')
        
        for key in tokens.keys():
            token = input('Por favor, escriba su {}.'.format(key))
            tokens[key] = token.strip()
            
        with open('oauth.json', 'w+') as file:
            file.write(json.dumps(tokens))
        
        # This checks if everything went fine
        with open('oauth.json', 'r') as file:
            oauth = json.load(file)
        if oauth == tokens:
            print('El archivo se ha creado correctamente.')
            
            
    def getUsers(self):
        
        mode = self.CONFIG['get_users_mode']
        
        if mode == 'file':
            try:
                with open('usuarios.txt', 'r') as file:
                    return [line.strip() for line in file]
            except:
                print('No existe el archivo usuarios.txt o no es correcto. Se procederá a usar el modo "terminal" para obtener los usuarios.')
                self.CONFIG['get_users_mode'] = 'terminal'
                return self.getUsers()                
        elif mode =='terminal':
            users = []
            user = input('Introduzca el primer nombre: ')
            if user != '':
                users.append(user)
            
            while user != '':
                user = input('Introduzca el siguiente nombre (pulse enter sin escribir nada para terminar): ')
                if user != '':
                    users.append(user)
            return users
        
        else:
            self.CONFIG['get_users_mode'] = 'terminal'
            return self.getUsers()
    
    
    def getConfigInfo(self):
        
        descriptions = {
                'use_last_ids': 'Con valor True, se recolectarán tweets a partir del último registrado por el programa.',
                'show_sleep': 'Con valor True uestra un aviso cada vez que el programa descanse para evitar llegar al límite de la API. AVISO: el volumen de avisos generado será, generalmente, alto.',
                'show_info': 'Con valor True muestra información relativa al error de tiempo (diferencia entre tiempo real de ejecución y tiempo estimado, y el tiempo total de ejecución del programa.',
                'show_tweet_quantity': 'Con valor True muestra información sobre los tweets recolectados, en cada ronda y en total, en cada iteración. AVISO: puede generar un volumen elevado de avisos.',
                'show_rate_limit': 'Muestra, al terminar la ejecución del programa, los límites de los "endpoints" de la API usados en el programa.',
                'get_users_mode': 'Puede tener dos valores:\narchivo - Los nombres de los usuarios se obtienen de un archivo llamado usuarios.txt, ubicado en el directorio base del programa.\nterminal - Se obtienen los usuarios desde la terminal. Se solicitan nombres al usuario hasta que se proporciona un "input" vacío, lo cual se puede conseguir pulsando enter sin escribir nada. Los nombres solicitados por el programa son los nombres de la cuentas (screen name).',
                'auto_export': 'Con valor True, exporta los tweets a excel al finalizar la ejecución del programa.',
                'export_mode': 'Puede tener dos valores:\nsplit - al exportar a excel se crean archivos para cada cuenta.\nunitary - al exportar a excel se crea un archivo para todos los usuarios usados en una sesión seleccionada por el usuario, con una división por hojas.'
                }
            
        for field in self.CONFIG.keys():
            print('>>> Campo: {}.\n>>> Descripción: {}\n'.format(field, descriptions[field]))
            
            
    def getConfig(self):
        
        print('{: <25}{: <15}'.format('Campo', 'Valor actual'))
        for field, value in self.CONFIG.items():
            print('{: <25}{: <15}'.format(field, str(value)))
            
            
    def setConfig(self, *args):
        
        for field in args:
            if field in self.CONFIG.keys():
                if field == 'get_users_mode':
                    if self.CONFIG[field] == 'terminal':
                        self.CONFIG[field] = 'archivo'
                    else:
                        self.CONFIG[field] = 'terminal'
                elif field == 'export_mode':
                    if self.CONFIG[field] == 'split':
                        self.CONFIG[field] = 'unitary'
                    else:
                        self.CONFIG[field] = 'split'
                else:
                    if self.CONFIG[field] == False:
                        self.CONFIG[field] = True
                    else:
                        self.CONFIG[field] = False
                print('{} ha sido cambiado correctamente a {}'.format(field, self.CONFIG[field]))
            else:
                print('{} no es un campo válido.'.format(field))
    
    
    def executeQueries(self):
        
        import twutils
        import json
        import time
        from datetime import datetime as t
        
        exec_time = time.clock()
        users = self.getUsers()
        
        if 'database' not in twutils.getFoldersInFolder():
            twutils.createFolder('./database/')
            
        if self.CONFIG['use_last_ids']:
            try:
                with open('ids.json','r') as ids_file:
                    ids = json.load(ids_file)
                    
                    last_tw_id = {user:ids[user] for user in users}
            
            except:
                last_tw_id = {user:None for user in users}
        else:
            last_tw_id = {user:None for user in users}

        now = t.now()
        folder_name = './database/{}-{}-{} {}h {}m/'.format(now.day, now.month, now.year, now.hour, now.minute)
        twutils.createFolder(folder_name)
        
        for user in users:
            with open('{}{}.json'.format(folder_name, user), 'w+') as file:
                file.write('{')
        
        counter = 0
        tweet_quantity = 0        
        test_time = time.clock()
        
        print('El programa se ha iniciado. No cierre el IDE, reinicie el núcleo o cierre los procesos asociados a este. Para pararlo se debe crear un archivo llamado "stop" (sin extensión de archivo) en la misma carpeta del archivo twitter_collector.py y esperar a que acabe en unos segundos.')
        
        while 'stop' not in twutils.getFilesInFolder('./'):
            
            row_tweet_quantity = 0
            
            for user in users:
                
                while not twutils.connected():
                    print('No se detecta conexión a internet. Volviendo a conectar en 15 segundos.')
                    time.sleep(15)
                    
                prev_time = time.clock()
                
                # Twitter statuses as dicts instead of Status objects
                raw_statuses = [status.__dict__ for status in self.api.GetUserTimeline(screen_name = user, count=200, since_id = last_tw_id[user])]
                
                # This determines if a comma is needed to be written in the json file
                comma = False
                
                if len(raw_statuses) > 0:
                    # We must filter the results to get only the fields we're interested in
                    statuses = twutils.filterStatuses(raw_statuses)
                    
                    tweet_quantity += len(statuses)
                    row_tweet_quantity += len(statuses)
                    
                    encoded_statuses = json.dumps(statuses, sort_keys=True)
            
                    with open('{}{}.json'.format(folder_name, user), 'a+') as file:
                        
                        if comma:
                            file.write(',')
                            
                        comma = True
                        
                        file.write(encoded_statuses[1:-1])
                else:
                    statuses = {}
                
                elapsed_time = time.clock() - prev_time
                
                if elapsed_time < 2:
                    if self.CONFIG['show_sleep']:
                        print('Durmiendo: {}'.format(2 - elapsed_time))
                    time.sleep(2 - elapsed_time)
                    
                counter += 1
                if len(statuses) > 0:
                    max_id = max([int(key) for key in statuses.keys()])
                    if last_tw_id[user] == None or max_id > last_tw_id[user]:
                        last_tw_id[user] = max_id
                
            if self.CONFIG['show_tweet_quantity']:
                print('Tweets recopilados en esta ronda: {}'.format(row_tweet_quantity))
                print('Tweets recopilados en total: {}\n'.format(tweet_quantity))
        
        for user in users:
            with open('{}{}.json'.format(folder_name, user), 'a') as file:
                file.write('}')
                    
        with open('ids.json', 'w+') as ids_file:
            
            try:
                ids = json.load(ids_file)
            except:
                ids = {}
            
            for user, last_id in last_tw_id.items():
                ids[user] = last_id
                
            ids_file.write(json.dumps(ids))
                
        test_time = time.clock() - test_time
        
        exec_time = time.clock() - exec_time
        
        print('El programa se ha detenido.')
        
        if self.CONFIG['show_info']:
            print('Tiempo de ejecución total: {}'.format(exec_time))
            print('Error de tiempo: {}'.format(test_time - (counter*2)))
            
        if self.CONFIG['show_rate_limit']:
            print('Solicitudes restantes (timeline): {}'.format(self.api.rate_limit.resources['statuses']['/statuses/user_timeline']['remaining']))
            print('Solicitudes restantes (rate limit status): {}'.format(self.api.rate_limit.resources['application']['/application/rate_limit_status']['remaining']))
            
        if self.CONFIG['auto_export']:
            try:
                self.excel_export(folder_name)
                print('Se han exportado a excel los tweets seleccionados.')
            except:
                print('Hubo un problema al exportar a excel los tweets.')


    def exportToExcel(self, folder_name):
        
        from openpyxl import Workbook
        from twutils import getFilesInFolder
        import json
        
        files = getFilesInFolder(folder_name)
        
        json_files = [file for file in files if file[-5:] == '.json']
        excel_files = [file[:-5] for file in files if file[-5:] == '.xlsx']
        
        if self.CONFIG['export_mode'] == 'split':
            
            for file in json_files:
                if file[:-5] in excel_files:
                    print('Ya se ha exportado {}.'.format(file))
                    continue
                
                with open(folder_name + file, 'r') as tw_file:
                    statuses = json.load(tw_file)              
                        
                if len(statuses) > 0:
                    wb = Workbook()
                    ws = wb.active
                    
                    ws.append(list(statuses[list(statuses.keys())[0]].keys()))
                    
                    for status in statuses.values():
                        ws.append(list(status.values()))
                        
                    wb.save('{}{}.xlsx'.format(folder_name, file[:-5]))
                    print('Se ha exportado {} en la ruta {}{}.xlsx.'.format(file, folder_name, file[:-5]))
                    
                else:
                    print('{} no contiene tweets. No se ha creado archivo excel relativo a este usuario.'.format(file))
                    
        elif self.CONFIG['export_mode'] == 'unitary':
            
            
            users_statuses = {}
            has_tweets = False
            
            for file in json_files:
                
                user = file[:-5]
                with open(folder_name + file, 'r') as tw_file:
                    tweets = json.load(tw_file)
                    if len(tweets) > 0:
                        has_tweets = True
                    if user not in users_statuses.keys():
                        users_statuses[user] = {}
                    users_statuses[user].update(tweets)
            
            if has_tweets:
                wb = Workbook()
                
                first_sheet = True
                for user, statuses in users_statuses.items():
                    if len(list(statuses.keys())) > 0:
                        
                        if first_sheet:
                            ws = wb.active
                            ws.title = user[1:]
                            first_sheet = False
                        else:
                            ws = wb.create_sheet(title=user[1:])
                            
                        ws.append(list(statuses[list(statuses.keys())[0]].keys()))
                        
                        for status in statuses.values():
                            ws.append(list(status.values()))
                        
                wb.save('./database/{}.xlsx'.format(folder_name[11:-1]))
                print('Se han exportado los archivos de la sesión {} en la ruta ./database/{}.xlsx.'.format(folder_name[11:-1], folder_name[11:-1]))
                
        else:
            self.CONFIG['export_mode'] = 'split'
            print('El modo seleccionado para este método es incorrecto. Se ha seleccionado automáticamente el modo "split". Pruebe a ejecutarlo de nuevo.')
            
    def exportAllToExcel(self):
        
        import twutils
        
        sessions = twutils.getFoldersInFolder('database')
        
        for session in sessions:
            if '{}.xlsx'.format(session) in twutils.getFilesInFolder('./database/'):
                print('Ya existe un archivo excel para la sesión {}.'.format(session))
            else:
                self.exportToExcel('./database/{}/'.format(session))
            
    def exportThisToExcel(self):
        
        from twutils import getFoldersInFolder
        
        sessions = getFoldersInFolder('database')
        
        print('Elija una sesión para exportar a excel:')
        
        count = 0
        for session in sessions:
            print('[{}] {}.'.format(count + 1, session))
            count += 1
            
        sel = input('Introduzca el número de la sesión:')
        
        try:
            sel = int(sel) - 1
            if sel >= 0:
                sel_session = sessions[sel]
                self.exportToExcel('./database/{}/'.format(sel_session))
        except Exception as E:
            print(E)
            self.exportThisToExcel()
            
        
